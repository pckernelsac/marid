from datetime import date, datetime, time

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.appointment import Appointment
from app.models.cashbox import CashMovement
from app.models.enums import CashMovementType, PaymentMethod
from app.models.patient import Patient
from app.models.treatment import Treatment
from app.schemas.report import (
    DayPoint,
    NamedAmount,
    NamedCount,
    ReportSummary,
)

STATUS_LABELS = {
    "pending": "Pendiente",
    "in_progress": "En proceso",
    "finished": "Finalizado",
    "scheduled": "Agendada",
    "confirmed": "Confirmada",
    "completed": "Completada",
    "cancelled": "Cancelada",
    "no_show": "No asistió",
}

METHOD_LABELS = {
    PaymentMethod.CASH: "Efectivo",
    PaymentMethod.CARD: "Tarjeta",
    PaymentMethod.TRANSFER: "Transferencia",
    PaymentMethod.YAPE: "Yape",
    PaymentMethod.PLIN: "Plin",
    PaymentMethod.OTHER: "Otro",
}


class ReportService:
    def __init__(self, db: Session) -> None:
        self.db = db

    def summary(self, date_from: date, date_to: date) -> ReportSummary:
        start_dt = datetime.combine(date_from, time.min)
        end_dt = datetime.combine(date_to, time.max)

        total_patients = self.db.execute(
            select(func.count(Patient.id))
        ).scalar_one()
        new_patients = self.db.execute(
            select(func.count(Patient.id)).where(
                Patient.created_at >= start_dt, Patient.created_at <= end_dt
            )
        ).scalar_one()

        # Treatments (by creation date)
        treat_filter = (
            Treatment.created_at >= start_dt,
            Treatment.created_at <= end_dt,
        )
        treatments_total = self.db.execute(
            select(func.count(Treatment.id)).where(*treat_filter)
        ).scalar_one()
        treat_status_rows = self.db.execute(
            select(Treatment.status, func.count(Treatment.id))
            .where(*treat_filter)
            .group_by(Treatment.status)
        ).all()
        treatments_by_status = [
            NamedCount(label=STATUS_LABELS.get(s.value, s.value), count=c)
            for s, c in treat_status_rows
        ]
        top_proc_rows = self.db.execute(
            select(Treatment.procedure, func.count(Treatment.id).label("c"))
            .where(*treat_filter)
            .group_by(Treatment.procedure)
            .order_by(func.count(Treatment.id).desc())
            .limit(5)
        ).all()
        top_procedures = [NamedCount(label=p, count=c) for p, c in top_proc_rows]

        # Appointments (by start)
        appt_filter = (
            Appointment.starts_at >= start_dt,
            Appointment.starts_at <= end_dt,
        )
        appointments_total = self.db.execute(
            select(func.count(Appointment.id)).where(*appt_filter)
        ).scalar_one()
        appt_status_rows = self.db.execute(
            select(Appointment.status, func.count(Appointment.id))
            .where(*appt_filter)
            .group_by(Appointment.status)
        ).all()
        appointments_by_status = [
            NamedCount(label=STATUS_LABELS.get(s.value, s.value), count=c)
            for s, c in appt_status_rows
        ]

        # Cash
        cash_filter = (
            CashMovement.movement_date >= date_from,
            CashMovement.movement_date <= date_to,
        )

        def _sum(mtype: CashMovementType) -> float:
            return float(
                self.db.execute(
                    select(func.coalesce(func.sum(CashMovement.amount), 0)).where(
                        *cash_filter, CashMovement.movement_type == mtype
                    )
                ).scalar_one()
            )

        income = _sum(CashMovementType.INCOME)
        expense = _sum(CashMovementType.EXPENSE)

        method_rows = self.db.execute(
            select(
                CashMovement.payment_method,
                func.coalesce(func.sum(CashMovement.amount), 0),
            )
            .where(*cash_filter, CashMovement.movement_type == CashMovementType.INCOME)
            .group_by(CashMovement.payment_method)
        ).all()
        income_by_method = [
            NamedAmount(label=METHOD_LABELS.get(m, m.value), amount=float(a))
            for m, a in method_rows
        ]

        series_rows = self.db.execute(
            select(
                CashMovement.movement_date,
                CashMovement.movement_type,
                func.coalesce(func.sum(CashMovement.amount), 0),
            )
            .where(*cash_filter)
            .group_by(CashMovement.movement_date, CashMovement.movement_type)
            .order_by(CashMovement.movement_date)
        ).all()
        series: dict[date, dict[str, float]] = {}
        for d, mtype, amount in series_rows:
            entry = series.setdefault(d, {"income": 0.0, "expense": 0.0})
            entry[mtype.value] = float(amount)
        revenue_series = [
            DayPoint(day=d, income=v["income"], expense=v["expense"])
            for d, v in sorted(series.items())
        ]

        return ReportSummary(
            date_from=date_from,
            date_to=date_to,
            total_patients=total_patients,
            new_patients=new_patients,
            treatments_total=treatments_total,
            treatments_by_status=treatments_by_status,
            top_procedures=top_procedures,
            appointments_total=appointments_total,
            appointments_by_status=appointments_by_status,
            total_income=income,
            total_expense=expense,
            balance=income - expense,
            income_by_method=income_by_method,
            revenue_series=revenue_series,
        )

    def export_excel(self, date_from: date, date_to: date) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        data = self.summary(date_from, date_to)
        wb = Workbook()
        header_fill = PatternFill("solid", fgColor="2563EB")
        header_font = Font(bold=True, color="FFFFFF")
        bold = Font(bold=True)

        def style_header(ws, row=1):
            for cell in ws[row]:
                cell.fill = header_fill
                cell.font = header_font

        # Sheet 1: Resumen
        ws = wb.active
        ws.title = "Resumen"
        ws.append(["Indicador", "Valor"])
        style_header(ws)
        rows = [
            ("Periodo", f"{date_from} a {date_to}"),
            ("Pacientes totales", data.total_patients),
            ("Pacientes nuevos", data.new_patients),
            ("Tratamientos", data.treatments_total),
            ("Citas", data.appointments_total),
            ("Ingresos", data.total_income),
            ("Egresos", data.total_expense),
            ("Balance", data.balance),
        ]
        for r in rows:
            ws.append(list(r))
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 22

        # Sheet 2: Ingresos por día
        ws2 = wb.create_sheet("Ingresos")
        ws2.append(["Fecha", "Ingresos", "Egresos"])
        style_header(ws2)
        for p in data.revenue_series:
            ws2.append([p.day.isoformat(), p.income, p.expense])
        ws2.append(["Total", data.total_income, data.total_expense])
        ws2[ws2.max_row][0].font = bold
        for col in ("A", "B", "C"):
            ws2.column_dimensions[col].width = 16

        # Sheet 3: Tratamientos
        ws3 = wb.create_sheet("Tratamientos")
        ws3.append(["Estado", "Cantidad"])
        style_header(ws3)
        for nc in data.treatments_by_status:
            ws3.append([nc.label, nc.count])
        ws3.append([])
        ws3.append(["Procedimiento", "Cantidad"])
        for nc in data.top_procedures:
            ws3.append([nc.label, nc.count])
        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 14

        # Sheet 4: Citas
        ws4 = wb.create_sheet("Citas")
        ws4.append(["Estado", "Cantidad"])
        style_header(ws4)
        for nc in data.appointments_by_status:
            ws4.append([nc.label, nc.count])
        ws4.column_dimensions["A"].width = 20
        ws4.column_dimensions["B"].width = 14

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
